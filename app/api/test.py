from flask import Blueprint, request, jsonify, current_app
from app.utils.auth import require_auth
from app.utils.database import safe_db_query
from .chat import ask_question

import openpyxl
import io
import logging
import uuid
import json

from app.utils.time_provider import get_current_datetime

test_bp = Blueprint("test", __name__)
logging.basicConfig(level=logging.INFO)

@test_bp.route("/test", methods=["POST"])
@require_auth
def test(**kwargs):
    user = kwargs.get("user")

    headers = {
        "Authorization": request.headers.get("Authorization"),  # teruskan token
        "Content-Type": "application/json"
    }

    # --- Validasi file ---
    if "file" not in request.files:
        return jsonify({"error": "File tidak ditemukan dalam request"}), 400

    file = request.files["file"]
    sheet_name = request.form.get("sheet_name", default="List")
    start_row = request.form.get("start_row", type=int, default=2)
    end_row = request.form.get("end_row", type=int, default=None)

    is_browse = request.form.get("is_browse", default=False, type=bool)
    is_company = request.form.get("is_company", default=True, type=bool)
    is_general = request.form.get("is_general", default=False, type=bool)

    # Generate single chat_id and session name
    chat_id = str(uuid.uuid4())
    datetime_str = get_current_datetime().strftime("%Y-%m-%d %H:%M:%S")
    session_name = f"QA Test {datetime_str}"

    # Create the chat session in database
    try:
        insert_chat_query = """
            INSERT INTO chats (id, user_id, subject, pinned, created_at, options)
            VALUES (%s, %s, %s, FALSE, CURRENT_TIMESTAMP, %s)
        """
        options = {"is_browse": is_browse, "is_company": is_company, "is_general": is_general}
        safe_db_query(insert_chat_query, (chat_id, user["user_id"], session_name, json.dumps(options)))
    except Exception as e:
        logging.error(f"Failed to create chat session: {e}")
        return jsonify({"error": "Gagal membuat sesi chat"}), 500

    try:
        # --- Baca workbook ---
        in_memory_file = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(in_memory_file, data_only=True)

        ws = wb[sheet_name]

        if not sheet_name:
            return jsonify({"error": "Parameter sheet_name wajib diisi"}), 400

        # --- Validasi sheet ---
        if sheet_name not in wb.sheetnames:
            return jsonify({"error": f"Sheet '{sheet_name}' tidak ditemukan"}), 400

        # --- Loop sesuai rentang baris ---
        results = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            number = row[0] if len(row) > 0 else None
            question = row[1] if len(row) > 1 else None
            expected_answer = row[2] if len(row) > 2 else None
            
            body = {
                "question": question,
                "chat_id": chat_id,  # Use the same chat_id for all questions
                "is_browse": is_browse,
                "is_company": is_company,
                "is_general": is_general
            }

             # --- INTERNAL CALL langsung ke fungsi ask_question ---
            with current_app.test_request_context( "/chats/ask", method="POST", json=body, headers=headers ): response = ask_question(user=user)  # panggil langsung fungsinya
            if isinstance(response, tuple):
                response_obj = response[0]  # ambil Response object
            else:
                response_obj = response

            response_data = response_obj.get_json()
            
            actual_answer = response_data.get("data", {}).get("answer", None)
            documents = response_data.get("data", {}).get("source_documents", [])
            # chat_id is now the same for all

            source = []
            for doc in documents:
                metadata = doc.get("metadata", {})
                if "Title" in metadata:
                    source.append(metadata.get("Title", None))
                elif "url" in metadata:
                    source.append(metadata.get("url", None))

            results.append({
                "number": number,
                "question": question,
                "expected_answer": expected_answer,
                "actual_answer_from_agent": actual_answer,
                "reference": source
            })

        # Clean up: delete the test chat session
        # try: 
        #     delete_chat_details_query = "DELETE FROM chat_details WHERE chat_id = %s"
        #     safe_db_query(delete_chat_details_query, (chat_id,))
            
        #     delete_chat_query = "DELETE FROM chats WHERE id = %s AND user_id = %s"
        #     safe_db_query(delete_chat_query, (chat_id, user["user_id"]))
        # except Exception as e:
        #     logging.error(f"test.chats_delete_test - query get chat error: {e}")
        #     return jsonify({"error": "Gagal memproses request"}), 500
        return jsonify(results)

    except Exception as e:
        logging.exception("Terjadi kesalahan saat membaca Excel")
        return jsonify({"error": str(e)}), 500
